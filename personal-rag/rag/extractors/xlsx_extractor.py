from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from rag.core.models import Page


def extract_xlsx(path: Path, doc_id: str) -> list[Page]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    pages: list[Page] = []
    for sheet in wb.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            line = " | ".join(str(v) for v in row if v is not None)
            if line:
                rows.append(line)
        pages.append(Page(doc_id, str(path), path.name, None, f"sheet {sheet.title}", "\n".join(rows), {"type": "xlsx", "sheet": sheet.title}))
    wb.close()
    return pages
